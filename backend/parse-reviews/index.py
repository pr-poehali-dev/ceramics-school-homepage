import base64
import io
import json
import os

import psycopg2
from openpyxl import load_workbook


def _cors() -> dict:
    return {
        'Access-Control-Allow-Origin': '*',
        'Access-Control-Allow-Methods': 'POST, OPTIONS',
        'Access-Control-Allow-Headers': 'Content-Type, X-Session-Token',
        'Access-Control-Max-Age': '86400',
    }


def handler(event: dict, context) -> dict:
    '''
    Разбирает Excel-файл (.xlsx) со списком отзывов и возвращает готовый JSON-массив.
    Ожидаемые колонки (в любом порядке, по названию заголовка в первой строке):
    Имя, Метка (необязательно), Дата, Оценка, Текст.
    Требует авторизации менеджера.
    Args: event с httpMethod, headers (X-Session-Token), body (fileData base64)
          context - объект с request_id
    Returns: HTTP-ответ со списком разобранных отзывов [{name, meta, date, rating, text}]
    '''
    method = event.get('httpMethod', 'GET')

    if method == 'OPTIONS':
        return {'statusCode': 200, 'headers': _cors(), 'body': ''}

    if method != 'POST':
        return {'statusCode': 405, 'headers': _cors(), 'body': json.dumps({'error': 'Method not allowed'})}

    headers = event.get('headers') or {}
    token = headers.get('X-Session-Token') or headers.get('x-session-token') or ''
    if not token:
        return {'statusCode': 401, 'headers': _cors(), 'body': json.dumps({'error': 'Требуется авторизация'})}

    conn = psycopg2.connect(os.environ['DATABASE_URL'])
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT manager_id FROM manager_sessions WHERE token = %s AND expires_at > NOW()",
            (token,),
        )
        if not cur.fetchone():
            return {'statusCode': 401, 'headers': _cors(), 'body': json.dumps({'error': 'Сессия истекла, войдите снова'})}
    finally:
        conn.close()

    body = json.loads(event.get('body') or '{}')
    file_data = body.get('fileData') or ''

    if not file_data:
        return {'statusCode': 400, 'headers': _cors(), 'body': json.dumps({'error': 'Файл не передан'})}

    if ',' in file_data:
        file_data = file_data.split(',', 1)[1]

    try:
        raw = base64.b64decode(file_data)
    except Exception:
        return {'statusCode': 400, 'headers': _cors(), 'body': json.dumps({'error': 'Некорректные данные файла'})}

    if len(raw) > 8 * 1024 * 1024:
        return {'statusCode': 400, 'headers': _cors(), 'body': json.dumps({'error': 'Файл больше 8МБ'})}

    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
    except Exception:
        return {'statusCode': 400, 'headers': _cors(), 'body': json.dumps({'error': 'Не удалось прочитать Excel-файл'})}

    if not rows:
        return {'statusCode': 400, 'headers': _cors(), 'body': json.dumps({'error': 'Файл пустой'})}

    header = [str(c).strip().lower() if c else '' for c in rows[0]]

    def find_col(*names):
        for i, h in enumerate(header):
            if h in names:
                return i
        return None

    col_name = find_col('имя', 'name', 'автор')
    col_meta = find_col('метка', 'meta', 'статус')
    col_date = find_col('дата', 'date')
    col_rating = find_col('оценка', 'rating', 'рейтинг')
    col_text = find_col('текст', 'text', 'отзыв')

    if col_name is None or col_text is None:
        return {
            'statusCode': 400,
            'headers': _cors(),
            'body': json.dumps({'error': 'В файле должны быть колонки «Имя» и «Текст»'}),
        }

    reviews = []
    for row in rows[1:]:
        if row is None or all(c is None for c in row):
            continue

        def get(idx):
            if idx is None or idx >= len(row):
                return None
            v = row[idx]
            return str(v).strip() if v is not None else None

        name = get(col_name)
        text = get(col_text)
        if not name or not text:
            continue

        rating_raw = get(col_rating)
        try:
            rating = int(float(rating_raw)) if rating_raw else 5
        except ValueError:
            rating = 5
        rating = max(1, min(5, rating))

        reviews.append({
            'name': name,
            'meta': get(col_meta) or '',
            'date': get(col_date) or '',
            'rating': rating,
            'text': text,
        })

    if not reviews:
        return {'statusCode': 400, 'headers': _cors(), 'body': json.dumps({'error': 'Не найдено ни одного отзыва в файле'})}

    return {
        'statusCode': 200,
        'headers': _cors(),
        'body': json.dumps({'ok': True, 'reviews': reviews, 'count': len(reviews)}, ensure_ascii=False),
    }
